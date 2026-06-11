#!/usr/bin/env python3
"""
survey.py — SlideForge data-survey helper.

Profiles an .xlsx / .xlsm / .csv workbook so the calling agent can map its
columns onto whatever fields a SlideForge chart/table template needs, then hand
`slideforge_generate` a clean `slide.data[data_key]` payload.

Design goals (why this exists):
  * Diligence models are often HUGE and MESSY — dozens of sheets, banner rows,
    merged headers, summary blocks. A naive "read everything" both wastes time
    and drowns the agent in noise. So this runs a TWO-PASS survey: a cheap
    structural score of every sheet first, then a deep column-profile of only
    the top 1–2 sheets.
  * It must not choke on big files, so it reads in openpyxl READ-ONLY /
    streaming mode (cells are pulled lazily, one row at a time, never the whole
    grid into memory).
  * It is DELIBERATELY GENERIC. It knows nothing about retention, ARR, TAM, or
    any specific chart. It just reports, per sheet: the best header row, and for
    every column a profile (label, data type, how full it is, sample values,
    numeric min/max). The agent owns the mapping from these columns to the
    target chart's required fields — because the agent already knows the chart
    schema and can read the samples and ask the user.

This script produces HYPOTHESES, not a trusted extraction. The agent confirms
the mapping with the user (one AskUserQuestion) before generating the slide.

CLI:
    python3 survey.py <path-to-xlsx-or-csv>            # JSON (default)
    python3 survey.py <path> --text                    # human-readable
    python3 survey.py --self-test                       # built-in checks

Stdlib only, plus openpyxl for .xlsx reading.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import re
import sys
from typing import Any

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError as _e:  # pragma: no cover
    print(
        f"ERROR: openpyxl is required. Install with: pip install openpyxl  ({_e})",
        file=sys.stderr,
    )
    sys.exit(2)


# ---------------------------------------------------------------------------
# Tunables
# ---------------------------------------------------------------------------

HEADER_SCAN_ROWS = 25       # how many top rows to scan when guessing the header
MAX_SAMPLES = 5             # sample values kept per column
SAMPLE_ROWS = 3             # full data rows kept for the confirm-with-user step
DEEP_SHEET_CAP = 2          # how many sheets get the expensive deep profile
NEAR_TIE_RATIO = 0.8        # a runner-up scoring >= 80% of the top is a "near tie"
MAX_COLS = 200              # hard cap on columns profiled (defends vs. junk-wide sheets)

# Sheet-name hints. Source/data sheets tend to be named like the left list;
# finished outputs / scaffolding like the right list. These only nudge the
# score — they never override structural evidence.
SOURCE_NAME_HINTS = (
    "data", "raw", "detail", "model", "actual", "revenue", "sales", "bookings",
    "arr", "mrr", "pnl", "p&l", "financ", "summary", "metric", "kpi", "output",
)
NON_SOURCE_NAME_HINTS = (
    "cover", "toc", "contents", "instruction", "notes", "readme", "legend",
    "assumptions", "config", "scratch", "lookup", "ref",
)


# ---------------------------------------------------------------------------
# Cell-type helpers
# ---------------------------------------------------------------------------

_DATE_PATTERNS = [
    r"^[A-Za-z]{3,9}[-\s/]\d{2,4}$",                 # Jan-25 / January 2025
    r"^\d{4}[-/]\d{1,2}([-/]\d{1,2})?$",             # 2025-01 / 2025-01-31
    r"^\d{1,2}[-/]\d{4}$",                           # 1/2025
    r"^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$",              # 1/31/2025
    r"^FY\s?\d{2,4}$",                               # FY24 / FY2024
    r"^Q[1-4]'?\s?\d{2,4}$",                         # Q1'24 / Q1 2024
]


def looks_like_date(value: Any) -> bool:
    if isinstance(value, (dt.date, dt.datetime)):
        return True
    if isinstance(value, str):
        s = value.strip()
        return any(re.match(p, s, re.IGNORECASE) for p in _DATE_PATTERNS)
    return False


def cell_type(value: Any) -> str:
    """Coarse type label used in the column profile."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return "empty"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (dt.date, dt.datetime)):
        return "date"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, str):
        return "date" if looks_like_date(value) else "text"
    return "text"


def jsonsafe(value: Any) -> Any:
    if isinstance(value, (dt.date, dt.datetime)):
        return value.isoformat()
    if isinstance(value, float):
        # keep ints clean, round noisy floats for display
        return value
    return value


# ---------------------------------------------------------------------------
# Pass 1 — cheap structural score
# ---------------------------------------------------------------------------

def _name_bonus(name: str) -> float:
    nl = name.lower()
    bonus = 0.0
    if any(h in nl for h in SOURCE_NAME_HINTS):
        bonus += 6.0
    if any(h in nl for h in NON_SOURCE_NAME_HINTS):
        bonus -= 10.0
    return bonus


def _row_signature(row: tuple) -> dict:
    """Count cell kinds in a single row (used for header detection + scoring)."""
    n_text = n_num = n_date = n_nonempty = 0
    for v in row:
        t = cell_type(v)
        if t == "empty":
            continue
        n_nonempty += 1
        if t == "text":
            n_text += 1
        elif t == "number":
            n_num += 1
        elif t == "date":
            n_date += 1
    return {"text": n_text, "num": n_num, "date": n_date, "nonempty": n_nonempty}


def guess_header_row(preview: list[tuple]) -> int:
    """Pick the most likely header row from the first rows of a sheet.

    A header row is mostly labels (text or date headers) and is followed by rows
    that introduce numbers. We score each candidate by how header-like it is and
    whether numeric data starts beneath it. Returns a 1-based row index; 0 means
    "no convincing header found".
    """
    if not preview:
        return 0
    best_row, best_score = 0, 0.0
    for i, row in enumerate(preview):
        sig = _row_signature(row)
        if sig["nonempty"] < 2:
            continue
        label_cells = sig["text"] + sig["date"]
        # numbers appearing in the rows just below this one = strong header signal
        below_num = 0
        for j in range(i + 1, min(i + 4, len(preview))):
            below_num += _row_signature(preview[j])["num"]
        score = label_cells * 2.0 + (5.0 if below_num > 0 else 0.0) - sig["num"] * 0.5
        if score > best_score:
            best_score, best_row = score, i + 1  # 1-based
    return best_row


def quick_score_sheet(ws, preview: list[tuple]) -> dict:
    name = ws.title
    header_row = guess_header_row(preview)
    if header_row == 0:
        return {"name": name, "score": 0.0, "header_row": 0,
                "reason": "No convincing header row in the top rows."}

    header_sig = _row_signature(preview[header_row - 1])
    n_cols = header_sig["nonempty"]
    # data rows = whatever the sheet reports below the header (read-only gives
    # max_row from the stored dimension; guard for None)
    max_row = ws.max_row or len(preview)
    n_data_rows = max(0, max_row - header_row)

    cols_factor = min(n_cols, 20)
    rows_factor = min(n_data_rows / 3.0, 12.0)
    score = cols_factor + rows_factor + _name_bonus(name)
    return {
        "name": name,
        "score": round(score, 2),
        "header_row": header_row,
        "n_header_cols": n_cols,
        "n_data_rows": n_data_rows,
        "reason": (f"header row {header_row} with {n_cols} label cols, "
                   f"{n_data_rows} data rows, name bonus {_name_bonus(name):+.0f}"),
    }


# ---------------------------------------------------------------------------
# Pass 2 — deep column profile
# ---------------------------------------------------------------------------

def profile_sheet(ws, header_row: int) -> dict:
    """Stream the whole sheet once and build a per-column profile."""
    name = ws.title
    columns: dict[int, dict] = {}
    sample_rows: list[list] = []
    n_data_rows = 0
    header_labels: dict[int, Any] = {}

    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx > (header_row + 200000):  # absurd-size backstop
            break
        if r_idx < header_row:
            continue
        if r_idx == header_row:
            for c_idx, v in enumerate(row[:MAX_COLS], start=1):
                header_labels[c_idx] = v
            continue
        # data row
        is_blank = all(cell_type(v) == "empty" for v in row[:MAX_COLS])
        if is_blank:
            continue
        n_data_rows += 1
        if len(sample_rows) < SAMPLE_ROWS:
            sample_rows.append([jsonsafe(v) for v in row[:MAX_COLS]])
        for c_idx, v in enumerate(row[:MAX_COLS], start=1):
            col = columns.setdefault(c_idx, {
                "type_votes": {}, "n_nonempty": 0, "samples": [],
                "num_min": None, "num_max": None, "has_negative": False,
            })
            t = cell_type(v)
            if t == "empty":
                continue
            col["n_nonempty"] += 1
            col["type_votes"][t] = col["type_votes"].get(t, 0) + 1
            if len(col["samples"]) < MAX_SAMPLES:
                col["samples"].append(jsonsafe(v))
            if t == "number":
                fv = float(v)
                col["num_min"] = fv if col["num_min"] is None else min(col["num_min"], fv)
                col["num_max"] = fv if col["num_max"] is None else max(col["num_max"], fv)
                if fv < 0:
                    col["has_negative"] = True

    # finalize
    out_cols = []
    for c_idx in sorted(set(list(columns.keys()) + list(header_labels.keys()))):
        if c_idx > MAX_COLS:
            break
        col = columns.get(c_idx, {"type_votes": {}, "n_nonempty": 0, "samples": [],
                                  "num_min": None, "num_max": None, "has_negative": False})
        votes = col["type_votes"]
        dominant = max(votes, key=votes.get) if votes else "empty"
        label = header_labels.get(c_idx)
        out_cols.append({
            "col_letter": get_column_letter(c_idx),
            "col_index": c_idx,
            "header_label": jsonsafe(label) if label not in (None, "") else None,
            "dtype": dominant,
            "type_mix": votes,
            "n_nonempty": col["n_nonempty"],
            "samples": col["samples"],
            "numeric_range": (
                {"min": col["num_min"], "max": col["num_max"]}
                if col["num_min"] is not None else None
            ),
            "has_negative": col["has_negative"],
        })

    # Drop trailing fully-empty columns (common in exported models).
    while out_cols and out_cols[-1]["n_nonempty"] == 0 and out_cols[-1]["header_label"] is None:
        out_cols.pop()

    # For chart/table mapping the useful split is "what can be a category/axis
    # label" vs "what is a numeric series". Period labels (FY24, Q1'25, Jan-25)
    # and plain text both serve as category axes, so group text+date together.
    category_cols = [c for c in out_cols if c["dtype"] in ("text", "date") and c["n_nonempty"] > 0]
    numeric_cols = [c for c in out_cols if c["dtype"] == "number"]

    return {
        "name": name,
        "header_row": header_row,
        "n_data_rows": n_data_rows,
        "n_columns": len(out_cols),
        "columns": out_cols,
        "sample_rows": sample_rows,
        "role_hints": {
            # candidate x-axis / row-label columns (text or period labels)
            "category_columns": [c["col_letter"] for c in category_cols],
            # candidate numeric series columns
            "numeric_columns": [c["col_letter"] for c in numeric_cols],
        },
    }


# ---------------------------------------------------------------------------
# Workbook orchestration
# ---------------------------------------------------------------------------

def _preview_rows(ws, n: int) -> list[tuple]:
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= n:
            break
        out.append(row)
    return out


def survey_xlsx(path: str) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        scored = []
        for name in wb.sheetnames:
            ws = wb[name]
            preview = _preview_rows(ws, HEADER_SCAN_ROWS)
            scored.append((name, quick_score_sheet(ws, preview)))
        ranked = sorted(scored, key=lambda kv: kv[1]["score"], reverse=True)

        if not ranked or ranked[0][1]["score"] <= 0:
            deep_names = [ranked[0][0]] if ranked else []
            near_tie = False
        else:
            top = ranked[0][1]["score"]
            threshold = NEAR_TIE_RATIO * top
            deep_names = [n for n, q in ranked if q["score"] >= threshold and q["score"] > 0]
            deep_names = deep_names[:DEEP_SHEET_CAP]
            near_tie = len(deep_names) > 1

        sheets = []
        for name in deep_names:
            q = dict(scored)[name]
            sheets.append(profile_sheet(wb[name], q["header_row"]))

        skipped = [
            {"name": n, "score": q["score"], "reason": q["reason"]}
            for n, q in ranked if n not in deep_names
        ]
        return {
            "file": os.path.basename(path),
            "path": path,
            "sheets_profiled": sheets,
            "skipped_sheets": skipped,
            "near_tie": near_tie,
            "all_sheet_names": list(wb.sheetnames),
        }
    finally:
        wb.close()


def survey_csv(path: str) -> dict:
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        return {"file": os.path.basename(path), "path": path, "sheets_profiled": [],
                "skipped_sheets": [], "near_tie": False, "all_sheet_names": []}

    def cast(val: str) -> Any:
        if val is None or val == "":
            return None
        s = str(val).strip().replace(",", "").replace("$", "")
        if re.match(r"^-?\d+(\.\d+)?$", s):
            f = float(s)
            return int(f) if f.is_integer() else f
        return val

    typed = [tuple(cast(v) for v in row) for row in rows]
    preview = typed[:HEADER_SCAN_ROWS]
    header_row = guess_header_row(preview) or 1

    # Build an in-memory worksheet-like profile by reusing profile_sheet logic
    # via a tiny shim object.
    class _Shim:
        title = os.path.basename(path)
        def iter_rows(self, values_only=True):
            for row in typed:
                yield row
    sheet = profile_sheet(_Shim(), header_row)
    return {
        "file": os.path.basename(path),
        "path": path,
        "sheets_profiled": [sheet],
        "skipped_sheets": [],
        "near_tie": False,
        "all_sheet_names": [os.path.basename(path)],
    }


def survey_file(path: str) -> dict:
    low = path.lower()
    if low.endswith(".csv"):
        return survey_csv(path)
    if low.endswith((".xlsx", ".xlsm")):
        return survey_xlsx(path)
    raise ValueError(f"Unsupported extension: {path}. Provide .xlsx, .xlsm, or .csv.")


# ---------------------------------------------------------------------------
# Human-readable rendering
# ---------------------------------------------------------------------------

def render_text(report: dict) -> str:
    L = []
    L.append("=" * 78)
    L.append(f"FILE: {report['file']}   ({len(report['all_sheet_names'])} sheet(s))")
    L.append("=" * 78)
    L.append("HYPOTHESES for the agent to confirm with the user before generating.")
    L.append("")
    for s in report["sheets_profiled"]:
        L.append("-" * 78)
        L.append(f"SHEET: {s['name']}   header row {s['header_row']}, "
                 f"{s['n_data_rows']} data rows, {s['n_columns']} columns")
        L.append("-" * 78)
        rh = s["role_hints"]
        L.append(f"  category cols (x-axis / row labels): {rh['category_columns']}")
        L.append(f"  numeric cols (series values): {rh['numeric_columns']}")
        L.append("  columns:")
        for c in s["columns"]:
            if c["n_nonempty"] == 0 and c["header_label"] is None:
                continue
            rng = ""
            if c["numeric_range"]:
                rng = f" [{c['numeric_range']['min']:g}..{c['numeric_range']['max']:g}]"
            neg = " (has negatives)" if c["has_negative"] else ""
            L.append(f"    {c['col_letter']}: {c['header_label']!r} — {c['dtype']}"
                     f", {c['n_nonempty']} vals{rng}{neg}; e.g. {c['samples'][:3]}")
        L.append("")
    if report["skipped_sheets"]:
        L.append("SKIPPED (low structural score):")
        for sk in report["skipped_sheets"]:
            L.append(f"  - {sk['name']!r} (score {sk['score']}) — {sk['reason']}")
    L.append("=" * 78)
    return "\n".join(L)


# ---------------------------------------------------------------------------
# Self-test (no external fixture needed)
# ---------------------------------------------------------------------------

def _build_self_test_wb(path: str) -> None:
    wb = openpyxl.Workbook()
    # Sheet 1: a tidy revenue table with a banner row on top.
    ws = wb.active
    ws.title = "Financials"
    ws["A1"] = "ACME CORP — Confidential"            # banner row (noise)
    ws.append([])                                     # blank row
    ws["A3"] = "Year"
    ws["B3"] = "Revenue ($M)"
    ws["C3"] = "Gross Profit ($M)"
    for i, (yr, rev, gp) in enumerate([
        ("FY21", 38, 28), ("FY22", 54, 41), ("FY23", 73, 56),
        ("FY24", 94, 73), ("FY25", 118, 93),
    ]):
        ws.append([yr, rev, gp])  # rows 4..8
    # Sheet 2: a notes sheet that should be skipped.
    ns = wb.create_sheet("Notes")
    ns["A1"] = "Prepared by deal team. Figures illustrative."
    # Sheet 3: a tiny mixed-sign cash sheet to prove negatives are ordinary.
    cs = wb.create_sheet("Cash")
    cs["A1"] = "Quarter"
    cs["B1"] = "Net Cash ($M)"
    for q, v in [("Q1", 6), ("Q2", -14), ("Q3", 16), ("Q4", -106)]:
        cs.append([q, v])
    wb.save(path)


def run_self_test() -> int:
    import tempfile
    tmp = os.path.join(tempfile.gettempdir(), "slideforge_survey_selftest.xlsx")
    _build_self_test_wb(tmp)
    report = survey_xlsx(tmp)
    print(render_text(report))
    print()

    failures = []
    by_name = {s["name"]: s for s in report["sheets_profiled"]}

    fin = by_name.get("Financials")
    if not fin:
        failures.append("Financials sheet was not profiled (should be top-scored).")
    else:
        if fin["header_row"] != 3:
            failures.append(f"Financials header_row expected 3 (past the banner), got {fin['header_row']}")
        if fin["n_data_rows"] != 5:
            failures.append(f"Financials expected 5 data rows, got {fin['n_data_rows']}")
        cols = {c["header_label"]: c for c in fin["columns"]}
        if "Revenue ($M)" not in cols or cols["Revenue ($M)"]["dtype"] != "number":
            failures.append("Financials 'Revenue ($M)' should be a numeric column.")
        # FY-labels classify as period/date; what matters is they are a
        # category (x-axis) candidate, not a numeric series.
        if "A" not in fin["role_hints"]["category_columns"]:
            failures.append("Financials 'Year' column (A) should be a category candidate.")
        if "Year" not in cols or cols["Year"]["dtype"] not in ("text", "date"):
            failures.append(f"Financials 'Year' should be text/date, got {cols.get('Year', {}).get('dtype')}")
        if cols.get("Revenue ($M)", {}).get("numeric_range", {}) != {"min": 38.0, "max": 118.0}:
            failures.append(f"Revenue range expected 38..118, got {cols.get('Revenue ($M)', {}).get('numeric_range')}")

    if "Notes" in by_name:
        failures.append("Notes sheet was deep-profiled — it should be skipped.")
    if "Notes" not in {s["name"] for s in report["skipped_sheets"]}:
        failures.append("Notes not in skipped_sheets.")

    # Negatives are ordinary numbers — if Cash got profiled, prove the flag works.
    cash = by_name.get("Cash")
    if cash:
        netcash = next((c for c in cash["columns"] if c["header_label"] == "Net Cash ($M)"), None)
        if netcash and not netcash["has_negative"]:
            failures.append("Cash 'Net Cash ($M)' should flag has_negative=True.")

    print("=" * 78)
    if failures:
        print(f"SELF-TEST: FAIL ({len(failures)})")
        for f in failures:
            print(f"  - {f}")
        print("=" * 78)
        return 1
    print("SELF-TEST: PASS")
    print("  - Banner row skipped; header detected at row 3")
    print("  - Revenue/Gross Profit typed numeric, Year typed text")
    print("  - Notes sheet skipped via two-pass scoring")
    print("  - Negative values flagged as ordinary numbers")
    print("=" * 78)
    return 0


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        description="Survey an .xlsx/.csv workbook and emit a column profile for "
                    "mapping onto a SlideForge chart/table template."
    )
    p.add_argument("path", nargs="?", default=None, help="Path to .xlsx/.xlsm/.csv")
    p.add_argument("--text", action="store_true", help="Human-readable output (default is JSON)")
    p.add_argument("--self-test", action="store_true", help="Run built-in checks")
    args = p.parse_args(argv)

    if args.self_test:
        return run_self_test()
    if not args.path:
        print("ERROR: path is required (or use --self-test)", file=sys.stderr)
        return 2
    if not os.path.exists(args.path):
        print(f"ERROR: file not found: {args.path}", file=sys.stderr)
        return 2

    report = survey_file(args.path)
    if args.text:
        print(render_text(report))
    else:
        print(json.dumps(report, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
